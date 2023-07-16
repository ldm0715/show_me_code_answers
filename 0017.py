# TODO 将students.xlsx中的信息转化为xml格式

from yattag import Doc, indent
from openpyxl import load_workbook

# 什么是XML文件？
"""
XML代表可扩展标记语言
XML 是一种标记语言，很像 HTML
XML 旨在存储和传输数据
XML 被设计为具有自描述性
XML 是 W3C 推荐标准
"""

wb = load_workbook("./output/students.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

# 添加标题信息
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<!-- \n     学生信息表 \n-->'

doc.asis(xml_header)
doc.asis(xml_schema)

# 创建主标签
with tag('root'):
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5):
        row = [cell.value for cell in row]
        with tag("student"):
            with tag("姓名"):
                text(row[0])
            with tag("score1"):
                text(row[1])
            with tag("score2"):
                text(row[2])
            with tag("score3"):
                text(row[3])

result = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=False
)

print(result)

with open("./output/students.xml", "w", encoding="UTF-8") as f:
    f.write(result)
