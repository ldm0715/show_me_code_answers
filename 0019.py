# TODO 将numbers.xlsx中的信息转化为xml格式

from yattag import Doc, indent
from openpyxl import load_workbook

wb = load_workbook("./output/numbers.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

# 添加标题信息
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<!--\n     数字信息  \n-->'

doc.asis(xml_header)
doc.asis(xml_schema)

# 创建主标签
with tag('root'):
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
        row = [cell.value for cell in row]
        with tag("data"):
            text(f"[{row[0]},{row[1]},{row[2]}]")

result = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=False
)

print(result)

with open("./output/numbers.xml", "w", encoding="UTF-8") as f:
    f.write(result)
