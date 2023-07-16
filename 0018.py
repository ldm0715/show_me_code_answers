# TODO 将city.xlsx中的信息转化为xml格式

from yattag import Doc, indent
from openpyxl import load_workbook

wb = load_workbook("./output/city.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()

# 添加标题信息
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<!--\n    城市信息  \n-->'

doc.asis(xml_header)
doc.asis(xml_schema)

# 创建主标签
with tag('root'):
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2):
        row = [cell.value for cell in row]
        with tag("city"):
            with tag("名称"):
                text(row[0])

result = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=False
)

print(result)

with open("./output/city.xml", "w", encoding="UTF-8") as f:
    f.write(result)
